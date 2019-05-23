# -*- coding: utf-8 -*-
""" 
RinoZero V3.0
Creada el dia: 01-06-2018, 19:54
Autor: Felipe Barahona B.

Libreria RinoZero, version 2.0 (V2.0)
Se estructura con cuatro clases:
    - Clase Constructor
    - Clase Entrenador
    - Clase Desarrollador
    - Clase Evaluador

"""
import openpyxl as opyx
import numpy as np
import math as mt
#import time as tm

#=============================================================================
#1. Clase Constructor. Estructura de Red
#=============================================================================

class Constructor():
    
    def __init__(self,neuronasXcapaLista):
        self.neuronasXcapaLista = neuronasXcapaLista
        
    def armar(self):
       
       capas = len(self.neuronasXcapaLista)
       neuronasXcapa = np.zeros((1,capas)) 
       neuronasEntrada = self.neuronasXcapaLista[0]
       neuronasSalida = self.neuronasXcapaLista[capas-1]
       for i in range(capas):
           neuronasXcapa[0][i] = self.neuronasXcapaLista[i]
       
       return capas, neuronasEntrada, neuronasSalida, neuronasXcapa
   
    
    def armarXLSX(self,nombreArchivo):
        archivo = opyx.load_workbook(str(self.nombreArchivo)+".xlsx")
        hoja_archivo = archivo.get_sheet_by_name("Hoja3")
        capas = hoja_archivo.max_column
        neuronasXcapa = np.zeros((1,capas))
        
        for i in range(capas):
            neuronasXcapa[0][i] = hoja_archivo.cell(row=1,column=i+1).value
            neuronasEntrada = int(neuronasXcapa[0][0])
            neuronasSalida = int(neuronasXcapa[0][capas-1])
        
        archivo.save(str(self.nombreArchivo)+".xlsx")       
        
        return capas, neuronasEntrada, neuronasSalida, neuronasXcapa

    
    def armarTXT(self,nombreArchivo):
        contador = 0
        archivo = open(str(self.nombreArchivo)+".txt")
        for linea in archivo:
            if contador == 0:
                lin = str(linea.rstrip())
                lin = lin.split()[1]
                capas = int(lin)
                neuronasXcapa = np.zeros((1,capas))
                contador = contador + 1
            
            elif contador == 1:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasEntrada = int(lin)
                neuronasXcapa[0][contador-1] = neuronasEntrada
                contador = contador + 1
            
            elif contador > 1 and contador < capas:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasXcapa[0][contador-1] = int(lin)
                contador = contador + 1
            
            elif contador == capas:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasSalida = int(lin)
                neuronasXcapa[0][contador-1] = neuronasSalida
                break
               
        return capas, neuronasEntrada, neuronasSalida, neuronasXcapa
    
    
#=============================================================================
#2. Clase Entrenador. Entrenamiento mediante descenso de Gradiente
#=============================================================================
#
# Procedimientos para reducir sobreajuste
# 1. Regularizacion
# 2. Drop-out (apagado)
# 3. Normalizacion por lotes (Batch normalization)
        
    
class Entrenador():
    
    def __init__(self,capas,neuronasXcapa,funcionActivacion,funcionCosto):
        self.capas = capas
        self.neuronasXcapa = neuronasXcapa
        self.funcionActivacion = funcionActivacion
        self.funcionCosto = funcionCosto
    
    """====================================================================
       Metodo para imprimir criterios de calculo
       ===================================================================="""
    
    def imprimir_metodo(self):
        print("Funcion Activacion: "+self.funcionActivacion)
        print("Funcion Costo: "+self.funcionCosto)
        

    """====================================================================
       Metodos para cargar datos 
       ===================================================================="""
    
    def cargarDatosXLSX(self,nombre_base_datos):
        
        baseDatos = opyx.load_workbook(nombre_base_datos+".xlsx")
        variablesEntrada = baseDatos.get_sheet_by_name("Hoja1")
        variablesSalida = baseDatos.get_sheet_by_name("Hoja2")
        
        numEntrenamientos = variablesEntrada.max_row
        neuronasEntrada = variablesEntrada.max_column
        neuronasSalida = variablesSalida.max_column
        
        x = np.zeros((numEntrenamientos,neuronasEntrada))
        y = np.zeros((numEntrenamientos,neuronasSalida))
        
        for i in range(numEntrenamientos):
            for j in range(neuronasEntrada):
                x[i][j] = variablesEntrada.cell(row=i+1,column=j+1).value
            for j in range(neuronasSalida):
                y[i][j] = variablesSalida.cell(row=i+1,column=j+1).value
        
        return x,y
    
    
    def cargarPesos(self,nombre_archivo_pesos):
        
        pesosActualizados = [0 for i in range(self.capas-1)]
        basePesosActualizados = opyx.load_workbook(nombre_archivo_pesos+".xlsx")
        
        for capa in range(self.capas-1):
            
            matrizPesos = basePesosActualizados.get_sheet_by_name("Hoja"+str(capa+1))
            filas = matrizPesos.max_row
            cols = matrizPesos.max_column
            pesosActualizados[capa] = np.zeros((filas,cols))
            
            for i in range(filas):
                for j in range(cols):
                    pesosActualizados[capa][i][j] = float(matrizPesos.cell(row=i+1,column=j+1).value)
        
        return pesosActualizados
    
                    
        
    
    """====================================================================
       Metodos para inicializar pesos
       ===================================================================="""

    def iniciarPesosAleatorios(self):
        P = [None for i in range(self.capas-1)]
        for capa in range(1,self.capas):
            P[capa-1] = np.random.random_sample((int(self.neuronasXcapa[0][capa-1])+
                                         1,int(self.neuronasXcapa[0][capa])))
        
        return P
    
    
    def iniciarPesosUno(self):
        P = [None for i in range(self.capas-1)]
        for capa in range(1,self.capas):
            P[capa-1] = np.ones((int(self.neuronasXcapa[0][capa-1])+
                                         1,int(self.neuronasXcapa[0][capa])))
        
        return P

    
    def iniciarPesosCero(self):
        P = [None for i in range(self.capas-1)]
        for capa in range(1,self.capas):
            P[capa-1] = np.zeros((int(self.neuronasXcapa[0][capa-1])+
                                         1,int(self.neuronasXcapa[0][capa])))
        
        return P   
 
    """====================================================================
       Metodos para calcular funciones de costo
       ===================================================================="""
    #funcion de costo logaritmico
    #-------------------------------
    def costo_Logaritmico(self,u,y):
        funcionCosto = 0
    
        for i in range(len(u[0][:])):
            funcionCostoParcial = -(y[0][i]*mt.log(u[0][i])+(1-y[0][i])*mt.log(1-u[0][i]))
            funcionCosto = funcionCosto + funcionCostoParcial
    
        return funcionCosto

    #funcion de costo cuadratica
    #------------------------------
    def costo_Cuadratico(self,u,y):
        funcionCosto = 0
    
        for i in range(len(u[0][:])):
            funcionCostoParcial = (1/2)*(u[0][i]-y[0][i])**2
            funcionCosto = funcionCosto + funcionCostoParcial
    
        return funcionCosto
    
    """====================================================================
       Metodo para calcular deltas
       ===================================================================="""
    
    def calcularDelta(self,pesos,delta_suce,derivadas):
           
        pesos = pesos[1:pesos.shape[0]]
        pesosXdelta = pesos.dot(delta_suce.transpose()).reshape((1,pesos.shape[0]))
        delta_prede = np.zeros((1,pesosXdelta.shape[1]))
        for i in range(delta_prede.shape[1]):
            delta_prede[0][i] = pesosXdelta.transpose()[i].dot(derivadas.transpose()[i])
    
        return delta_prede

    """====================================================================
       Metodo para propagar adelante
       ===================================================================="""
    
    def propagarAdelante(self,variables_entrada,P):
        
        u= [None for i in range(self.capas)]
        u[0]= variables_entrada.reshape(1,int(self.neuronasXcapa[0][0]))
        
        for capa in range(1,self.capas):
            u[capa]= np.zeros((1,int(self.neuronasXcapa[0][capa])))
            
            if self.funcionActivacion == "sigmoide":
                u[capa]= funcionSigmoide(u[capa-1],P[capa-1])
            
            elif self.funcionActivacion == "RELU":

                if capa == self.capas-1:
                    u[capa]=funcionSigmoide(u[capa-1],P[capa-1])
                else:
                    u[capa]=funcionRELU(u[capa-1],P[capa-1])
            
            elif self.funcionActivacion == "leakyRELU":

                if capa == self.capas-1:
                    u[capa]=funcionSigmoide(u[capa-1],P[capa-1])

                else:
                    u[capa]=funcionLeakyRELU(u[capa-1],P[capa-1])
       
        return u
    
    """====================================================================
       Metodo para calcular derivadas
       ===================================================================="""
       
    def calcularDerivadas(self,u):
        
        du = [None for i in range(self.capas)]
        for i in range(self.capas):
            if i>0 and i<self.capas-1:
                du[i]=np.zeros((1,int(self.neuronasXcapa[0][i])))
        
        if self.funcionActivacion == "sigmoide":
            for capa in range(1,self.capas-1):
                du[capa] = u[capa]*(1-u[capa])
        
        elif self.funcionActivacion == "RELU":
            for capa in range(1,self.capas-1):
                for neurona in range(int(self.neuronasXcapa[0][capa])):
                    if u[capa][0][neurona] == 0:

                        if capa == self.capas-1:
                            du[capa][0][neurona] = u[capa][0][neurona]*(1-
                              u[capa][0][neurona])
                        else:
                            du[capa][0][neurona]=0
                        
                    elif u[capa][0][neurona] > 0:

                        if capa == self.capas-1:
                            du[capa][0][neurona] = u[capa][0][neurona]*(1-
                              u[capa][0][neurona])
                        else:
                            du[capa][0][neurona] = 1
        
        elif self.funcionActivacion == "leakyRELU":
            for capa in range(1,self.capas-1):
                for neurona in range(int(self.neuronasXcapa[0][capa])):
                    if u[capa][0][neurona] > 0:

                        if capa == self.capas-1:
                            du[capa][0][neurona] = u[capa][0][neurona]*(1-
                              u[capa][0][neurona])
                        else:
                            du[capa][0][neurona] == 1
                        
                    else:
                                        
                        if capa == self.capas-1:
                            du[capa][0][neurona] = u[capa][0][neurona]*(1-
                              u[capa][0][neurona])
                        else:
                            du[capa][0][neurona]=.1
                        
        return du
    
    """===================================================================
       Metodo para propagar atras
       ==================================================================="""
    
    """===================================================================
       Metodo para actualizar pesos
       ==================================================================="""
    
    """===================================================================
       Metodo para calcular costo en descenso de gradiente
       ==================================================================="""
       
    def calcularCostoDescGrad(self,x,y,pesos_iniciales):
        costo = 0
        for i in range(x.shape[0]):
            activaciones = self.propagarAdelante(x[i][:],pesos_iniciales)

            if self.funcionCosto == "logaritmica":
                costo_parcial = self.costo_Logaritmico(activaciones[self.capas-1],
                                                       y[i].reshape((1,y[i].shape[0])))
            elif self.funcionCosto == "cuadratica":
                costo_parcial = self.costo_Cuadratico(activaciones[self.capas-1],
                                                      y[i].reshape((1,y[i].shape[0])))
            
            costo = costo + costo_parcial
        costo = costo/(x.shape[0])
            
        return(costo)
  
    """===================================================================
       Metodos para efectuar entrenamiento de la red
       ==================================================================="""
    
    """===================================================================
       Metodo del descenso de gradiente
       ==================================================================="""
    
    def descensoGradiente(self,alpha,tol,x,y,pesos_iniciales):
        
        costo = 10
        contador = 0
        
        #Inicializar deltas
        delta = [None for i in range(self.capas)]
        pesos_actualizados = [None for i in range(len(pesos_iniciales))]
        print("RinoZero esta iniciando el entrenamiento...")
        
        while costo > tol:
            Delta = self.iniciarPesosCero()
            
            for i in range(x.shape[0]):
                #propagar adelante
                activaciones = self.propagarAdelante(x[i][:],pesos_iniciales)
                derivadas = self.calcularDerivadas(activaciones)
                delta[self.capas-1] = activaciones[len(activaciones)-1]-y[i]
                
                #propagar atras
                for capa in reversed(range(1,self.capas-1)):
                    delta[capa] = self.calcularDelta(pesos_iniciales[capa],delta[capa+1],derivadas[capa])
                
                for intercapa in range(self.capas-1):
                    pesos_actualizados[intercapa] = np.hstack((np.ones((1,1)),
                                      activaciones[intercapa])).transpose().dot(delta[intercapa+1])
                    Delta[intercapa] = Delta[intercapa]+pesos_actualizados[intercapa]
            
            #actualizaciones
            for intercapa in range(self.capas-1):
                pesos_iniciales[intercapa]=pesos_iniciales[intercapa]-alpha*(1/x.shape[0])*Delta[intercapa]
        
            #calculo del costo total
            #=======================================
            costo = self.calcularCostoDescGrad(x,y,pesos_iniciales)
                        
            contador +=1
            self.imprimir_res(contador,costo)
            #========================================
            
        print("RinoZero termino el entrenamiento...")
        return pesos_iniciales
    
    """===================================================================
       Metodo del descenso de gradiente estocastico
       ==================================================================="""    
    
    def descensoGradienteEstocastico(self,alpha,tol,x,y,pesos_iniciales):
        
        costo = 10
        contador = 0
        
        delta = [None for i in range(self.capas)]
        pesos_actualizados = [None for i in range(len(pesos_iniciales))]
        print("RinoZero esta iniciando el entrenamiento...")
        
        while costo > tol:
            Delta = self.iniciarPesosCero()
            
            xy = np.concatenate((x,y),axis=1)
            np.random.shuffle(xy)
            XY = np.hsplit(xy,[x.shape[1]])
            
            x=XY[0]
            y=XY[1]
                        
            for i in range(x.shape[0]):

                activaciones = self.propagarAdelante(x[i][:],pesos_iniciales)
                derivadas = self.calcularDerivadas(activaciones)
                delta[self.capas-1] = activaciones[len(activaciones)-1]-y[i]
                
                for capa in reversed(range(1,self.capas-1)):
                    delta[capa] = self.calcularDelta(pesos_iniciales[capa],delta[capa+1],derivadas[capa])
                
                for intercapa in range(self.capas-1):
                    pesos_actualizados[intercapa] = np.hstack((np.ones((1,1)),
                                      activaciones[intercapa])).transpose().dot(delta[intercapa+1])
                    Delta[intercapa] = Delta[intercapa]+pesos_actualizados[intercapa]

                for intercapa in range(self.capas-1):
                    pesos_iniciales[intercapa]=pesos_iniciales[intercapa]-alpha*Delta[intercapa]
            
            costo = self.calcularCostoDescGrad(x,y,pesos_iniciales)

            contador +=1
            self.imprimir_res(contador,costo)
        
        print("RinoZero termino el entrenamiento...")
        return pesos_iniciales    
        
    """==================================================================
       Metodo del descenso de gradiente por numero de iteraciones
       =================================================================="""
    def descensoGradiente2(self,alpha,niter,x,y,pesos_iniciales):
        
        costo = 10
        contador = 0
        
        #Inicializar deltas
        delta = [None for i in range(self.capas)]
        pesos_actualizados = [None for i in range(len(pesos_iniciales))]
        print("RinoZero esta iniciando el entrenamiento...")
        
        for k in range(niter):
            
            Delta = self.iniciarPesosCero()
            
            for i in range(x.shape[0]):
                #propagar adelante
                activaciones = self.propagarAdelante(x[i][:],pesos_iniciales)
                derivadas = self.calcularDerivadas(activaciones)
                delta[self.capas-1] = activaciones[len(activaciones)-1]-y[i]
                
                #propagar atras
                for capa in reversed(range(1,self.capas-1)):
                    delta[capa] = self.calcularDelta(pesos_iniciales[capa],delta[capa+1],derivadas[capa])
                
                for intercapa in range(self.capas-1):
                    pesos_actualizados[intercapa] = np.hstack((np.ones((1,1)),
                                      activaciones[intercapa])).transpose().dot(delta[intercapa+1])
                    Delta[intercapa] = Delta[intercapa]+pesos_actualizados[intercapa]
            
            #actualizaciones
            for intercapa in range(self.capas-1):
                pesos_iniciales[intercapa]=pesos_iniciales[intercapa]-alpha*(1/x.shape[0])*Delta[intercapa]
        
            #calculo del costo total
            #=======================================
            costo = self.calcularCostoDescGrad(x,y,pesos_iniciales)
                        
            contador +=1
            self.imprimir_res(contador,costo)
            #========================================
            
        print("RinoZero termino el entrenamiento...")
        
#        err_residual = open("C:/Users/Usuario/FAAM-Terreno/rinozero-flujo/1_ce/error_residual"+
#                            tm.strftime("%H_%M")+".txt","w")
#        err_residual.write(tm.strftime("%H:%M:%S")+"  "+str(niter)+"  "+str(costo))
#        err_residual.close()       
        
        return pesos_iniciales       
       

    """==================================================================
       Metodo para imprimir resultados de entrenamiento
       =================================================================="""
       
    def imprimir_res(self,contador,costo):
        lote = 1
        iteracion = contador
        epoca = lote*iteracion
        
        print("Iter: ",contador," Epoca: ",epoca," Funcion Costo: ",costo)
        
    """==================================================================
       Metodos para almacenar resultados
       =================================================================="""
    
    def registrarPesosXLSX(self,pesos,nombreArchivo):
        
        res = opyx.Workbook()
        hojas = [None for i in range(len(pesos))]
        
        for i in range(len(pesos)):
            hojas[i] = res.create_sheet("Hoja"+str(i+1))
        
        for hoja in range(len(hojas)):
            for i in range(pesos[hoja].shape[0]):
                for j in range(pesos[hoja].shape[1]):
                    hojas[hoja].cell(row=i+1,column=j+1).value = pesos[hoja][i][j]
        
        res.save(str(nombreArchivo+".xlsx"))
        
    
    def registrarPesosCSV(self,pesos,nombreArchivo):
        
        pass
    
    
    def registrarPesosTXT(self,pesos,nombreArchivo):
        
        pass
    
    def registrarErrorResidual(self,costo):
        error_residual = costo
        return error_residual
        
        
#=============================================================================
#Funciones activacion
#=============================================================================

def funcionSigmoide(x,P):
    x= np.hstack((np.ones((1,1)),x))
    u= x.dot(P)
    u= 1/(1+np.exp(-u))
    
    return u


def funcionRELU(x,P):
    x = np.hstack((np.ones((1,1)),x))
    u = x.dot(P)
    u = np.maximum(0,u)
    
    return u     


def funcionLeakyRELU(x,P):
    x = np.hstack((np.ones((1,1)),x))
    u = x.dot(P)
    np.maximum(u,0.01*u)
   
    return u


def funcionTANH(x,P):
    x = np.hstack((np.ones((1,1)),x))
    u = x.dot(P)
    
    for i in range(u.shape[1]):
        pass
    
    return u

#=============================================================================
#Funcion de propagaacion adelante
#=============================================================================
    
def prop_Adelante(entrada,pesos,capas,neuronasXcapa,funcionActivacion):
    pass



 

#=============================================================================
#3. Clase Desarrollador
#=============================================================================
class Desarrollador():
    def __init__(self,tipo_soporte):
        self.tipo_soporte = tipo_soporte
        
    
    def imprimir_soporte(self):
        print("El soporte de resultados es: ",self.tipo_soporte)



#=============================================================================
#4. Clase Evaluador
#=============================================================================
class Evaluador():
    def __init__(self,nombre_proyecto):
        self.nombre_proyecto = nombre_proyecto
    

    """==================================================================
       Metodo para leer archivos de pesos
       =================================================================="""    
    
    def leerPesosXLSX(self,nombreArchivo,numeroCapas):
        
        pesosCalculados=opyx.load_workbook(nombreArchivo+".xlsx")
        pesos = [0 for i in range(numeroCapas-1)]
        for i in range(len(pesos)):
            pesosIntercapa = pesosCalculados.get_sheet_by_name("Hoja"+str(i+1))
            matrizIntercapa = np.zeros((pesosIntercapa.max_row,pesosIntercapa.max_column))
            
            for j in range(pesosIntercapa.max_row):
                for k in range(pesosIntercapa.max_column):
                    matrizIntercapa[j][k] = pesosIntercapa.cell(row=j+1,column=k+1).value
            
            pesos[i] = matrizIntercapa
        pesosCalculados.save(nombreArchivo+".xlsx")
        
        return pesos

    """==================================================================
       Metodo para leer muestras del conjunto de prueba
       ==================================================================""" 
    
    def leerMuestrasXLSX(self,nombreArchivo):
        
        muestrasXLSX = opyx.load_workbook(nombreArchivo+".xlsx")
        muestras = muestrasXLSX.get_sheet_by_name("Hoja1")

        matrizMuestras = np.zeros((muestras.max_row,muestras.max_column))

        for i in range(muestras.max_row):
            for j in range(muestras.max_column):
                matrizMuestras[i][j] = muestras.cell(row=i+1,column=j+1).value
        
        return matrizMuestras	

    """==================================================================
       Metodo para leer etiquetas de las muestras del conjunto de prueba
       ==================================================================""" 

    def leerEtiquetasXLSX(self,nombreArchivo):
        
        muestrasXLSX = opyx.load_workbook(nombreArchivo+".xlsx")
        etiquetas = muestrasXLSX.get_sheet_by_name("Hoja2")
        
        matrizEtiquetas = np.zeros((etiquetas.max_row,etiquetas.max_column))

        for i in range(etiquetas.max_row):
            for j in range(etiquetas.max_column):
                matrizEtiquetas[i][j] = etiquetas.cell(row=i+1,column=j+1).value
        
        return matrizEtiquetas 	
    
    """==================================================================
       Metodo para determinar los numeros codigo de las muestras del conjunto de prueba
       =================================================================="""     
      
    def codif_EtiquetasXLSX(self,matrizEtiquetas):
        
        codigosCategorias = [0 for i in range(matrizEtiquetas.shape[0])]
        
        for i in range(matrizEtiquetas.shape[0]):
            muestra = matrizEtiquetas[i][:].reshape(1,matrizEtiquetas.shape[1])
            
            maxVal = np.amax(muestra)
            indice = 0
            
            for j in range(matrizEtiquetas.shape[1]):
                if maxVal - muestra[0][j] == 0:
                    indice = j
                    break
                else:
                    indice += 1
                
            codigosCategorias[i] = indice
        
        return codigosCategorias
    
    """==================================================================
       Metodo para ejecutar clasificacion en el conjunto de pruueba
       =================================================================="""         

    def ejecutarClasificacion(self,pesos,conjunto_muestras):
        """debe remplazarse por la funcion propagarAdelante"""
        pass
    