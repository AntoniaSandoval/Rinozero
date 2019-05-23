# -*- coding: utf-8 -*-
""" 
RinoZero V3.0
Creada el dia: 30-03-2018, 11:52
Autor: Felipe Barahona B.

Libreria RinoZero, version 3.0 (V3.0)
Redes convolucionales
Se estructura con cuatro clases:
    - Clase Constructor
    - Clase Entrenador
    - Clase Desarrollador
    - Clase Evaluador

"""
import openpyxl as opyx
import numpy as np


#============================================================================
#1. Clase Constructor
#============================================================================

class Constructor():
    
    def __init__(self,neuronasXcapaLista):
        self.neuronasXcapaLista = neuronasXcapaLista
    
    
    def armar(self):
        
        capas = len(self.neuronasXcapaLista)
        neuronasXcapa = np.zeros((1,capas))
        neuronasEntrada = self.neuronasXcapaLista[0]
        neuronasSalida = self.neuronasXcapaLista[capas-1]
        for i in range(capas):
            neuronasXcapa[0][i] = self.neuronasXcapaLista[i]
        
        return capas,neuronasEntrada, neuronasSalida, neuronasXcapa
    
    def armarXLSX(self,nombreArchivo):
        archivo = opyx.load_workbook(str(self.nombreArchivo)+".xlsx")
        hoja_archivo = archivo.get_sheet_by_name("Hoja3")
        capas = hoja_archivo.max_column
        neuronasXcapa = np.zeros((1,capas))
        
        for i in range(capas):
            neuronasXcapa[0][i] = hoja_archivo.cell(row=1,column=i+1).value
            neuronasEntrada = int(neuronasXcapa[0][0])
            neuronasSalida = int(neuronasXcapa[0][capas-1])
        
        archivo.save(str(self.nombreArchivo)+".xlsx")       
        
        return capas, neuronasEntrada, neuronasSalida, neuronasXcapa

    def armarTXT(self,nombreArchivo):
        contador = 0
        archivo = open(str(self.nombreArchivo)+".txt")
        for linea in archivo:
            if contador == 0:
                lin = str(linea.rstrip())
                lin = lin.split()[1]
                capas = int(lin)
                neuronasXcapa = np.zeros((1,capas))
                contador = contador + 1
            
            elif contador == 1:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasEntrada = int(lin)
                neuronasXcapa[0][contador-1] = neuronasEntrada
                contador = contador + 1
            
            elif contador > 1 and contador < capas:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasXcapa[0][contador-1] = int(lin)
                contador = contador + 1
            
            elif contador == capas:
                lin = str(linea.rstrip())
                lin = lin.split()[3]
                neuronasSalida = int(lin)
                neuronasXcapa[0][contador-1] = neuronasSalida
                break
               
        return capas, neuronasEntrada, neuronasSalida, neuronasXcapa    

#=============================================================================
#2. Clase Entrenador
#Incopora - operaciones de convolucion, pooling
#Convolucion: --> convolucion, padding, stride
#=============================================================================



        