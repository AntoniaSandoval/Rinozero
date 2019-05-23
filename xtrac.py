# -*- coding: utf-8 -*-
"""
Created on Fri Jun 15 15:15:07 2018

@author: Usuario
Clase xtrac

extrae informacion y genera estructuras de datos
para el procesamiento con los modulos de calculo

"""
import numpy as np
import openpyxl as opyx
import re
import csv
import xlrd
import urllib
import socket

class DiscoDuro:
    
    def __init__(self,nombreArchivo):
        
        self.nombreArchivo = nombreArchivo
        
    
    def obtenerExtension(self):    
        
        cadenaNombre = self.nombreArchivo.split(".")
        extension = cadenaNombre[len(cadenaNombre)-1]
        
        return extension
    

    def extraerDataCSV(self,dimDATA,dimETIQ):
        
        data = []
        indice = 0
        
        for linea in open(self.nombreArchivo,"r").readlines():
            
            data.append(linea)
            data[indice] = data[indice].split(";")
            data[indice][dimDATA+dimETIQ-1] = re.sub("\n","",
                data[indice][dimDATA+dimETIQ-1])
            indice +=1
        
        return data
    

    def extraerDataXLSX(self):
        
        obj_xlsx = opyx.load_workbook(self.nombreArchivo)
        hoj_xlsx = obj_xlsx.get_sheet_by_name("Hoja1")
        data = [None for i in range(hoj_xlsx.max_row)]
        
        for i in range(hoj_xlsx.max_row):
            data[i] = [None for j in range(hoj_xlsx.max_column)]
        
        for i in range(hoj_xlsx.max_row):
            for j in range(hoj_xlsx.max_column):
                data[i][j] = hoj_xlsx.cell(row=i+1,column=j+1).value
        
        obj_xlsx.save(self.nombreArchivo)
        
        return data
    

    def extraerDataXLS(self):
        
        obj_xls = xlrd.open_workbook(self.nombreArchivo)
        hoj_xls = obj_xls.sheet_by_index(0)
        data = [None for i in range(hoj_xls.nrows)]
        
        for i in range(hoj_xls.nrows):
            data[i] = [None for j in range(hoj_xls.ncols)]
        
        for i in range(hoj_xls.nrows):
            for j in range(hoj_xls.ncols):
                data[i][j] = hoj_xls.cell_value(rowx=i,colx=j)
        
        return data
    
        
    def derivarFormato(self,dimDATA,dimETIQ):
        
        extension = self.obtenerExtension()
        
        if extension == "txt":
            data = self.extraerDataCSV()
        elif extension == "csv":
            data = self.extraerDataCSV(dimDATA,dimETIQ)
        elif extension == "xlsx":
            data = self.extraerDataXLSX()
        elif extension == "xls":
            data = self.extraerDataXLS()
        
        return data
    

    def separarData(self,data,dimDATA,dimETIQ):
        
        x = []
        y = []
        
        for i in range(len(data)):
        
            x.append(data[i][0:dimDATA])
            y.append(data[i][dimDATA:dimDATA+dimETIQ])
        
        return x,y
        
    
    def cargarAtributos(self,nombreArchivoAtributo):
        
        atributo = open(nombreArchivoAtributo,"r").readlines()
        atributo = atributo[0].split(";")
        
        return atributo
    
    
    def chequearExtraccion(self,x,y,atributo,nombreArchivo,extension):
        
        pass
    

    def armarData(self,dimDATA,dimETIQ):
        
        data = self.derivarFormato(dimDATA,dimETIQ)
        sep_data = self.separarData(data,dimDATA,dimETIQ)
        x = sep_data[0]
        y = sep_data[1]
        
        return (x,y)


    def transformarNumpy(self,arreglo):
        
        numpy_arreglo = np.zeros((len(arreglo),len(arreglo[0])))
        for i in range(len(arreglo)):
            for j in range(len(arreglo[0])):
                try:
                    numpy_arreglo[i][j] = float(arreglo[i][j])
                except:
                    print("Existen valores no transformables a flotante")
                    break
                        
        return numpy_arreglo
    
    
    def transformarFlotante(self,arreglo,cols):
        
        for k in range(len(cols)):
            for i in range(len(arreglo)):
                for j in range(len(arreglo[0])):
                    if j == cols[k]:
                        arreglo[i][j]=float(arreglo[i][j])
        
        return arreglo
                

    def extraerPesosCSV(self,nombreArchivoPesos,neuronasPreSinap,neuronasPostSinap):
        
        archivo = open(nombreArchivoPesos,'r')
        contenido = archivo.readlines()
               
        for i in range(neuronasPreSinap+1):
            contenido[i] = contenido[i].split(';')
            contenido[i][len(contenido[i])-1] = contenido[i][len(contenido[i])-1].replace('\n','')
                        
        pesos = np.zeros((neuronasPreSinap+1,neuronasPostSinap))

        for i in range(neuronasPreSinap+1):
            for j in range(neuronasPostSinap):

                pesos[i][j] = float(contenido[i][j])
        
        return pesos
        
     
    def cargarPesosCSV(self, listaArchivosPesos, estructuraRed):
        
        pesos = []
               
        for i in range(len(listaArchivosPesos)):
            pesos.append(self.extraerPesosCSV(listaArchivosPesos[i],estructuraRed[i],estructuraRed[i+1]))
        
        
        return pesos
 
    

#class Internet:
#    
#    def __init__(self,nombreURL):
#        nombreURL = self.nombreURL
#
#class Periferico:
#    
#    def __init__(self,nombrePuerto):
#        nombrePuerto = self.nombrePuerto
    
class Preprocesador:
    
    def __init__(self,data):
        self.data = data
        
    def normalizar(self):
        pass
    